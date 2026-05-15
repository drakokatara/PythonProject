import json
import holidays
from datetime import date
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from .models import Employee, Attendance
from .forms import EmployeeForm

def get_greek_holidays():
    gr_holidays = holidays.Greece(years=date.today().year)
    events = [{
        'title': f"🎉 {name}",
        'start': d.strftime('%Y-%m-%d'),
        'color': '#ffc107',
        'textColor': '#000',
        'allDay': True
    } for d, name in gr_holidays.items()]
    return gr_holidays, events

def manage_employees(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Ο υπάλληλος προστέθηκε!")
            return redirect('manage_employees')
        messages.error(request, "❌ Σφάλμα στην εγγραφή. Ελέγξτε τα στοιχεία.")

    form = EmployeeForm()
    today = date.today()

    employees_qs = Employee.objects.prefetch_related('attendance_set').all()
    gr_holidays, holiday_events = get_greek_holidays()

    data = []
    colors = {
        'OFFICE': '#10b981',
        'REMOTE': '#0ea5e9',
        'LEAVE': '#f59e0b',
        'SICK': '#f59e0b'
    }

    total_debt_accumulator = 0

    for emp in employees_qs:
        report = emp.get_monthly_report()
        total_debt_accumulator += report['debt']

        today_att = emp.attendance_set.filter(date=today).first()
        today_status = today_att.work_type if today_att else 'NONE'

        events_list = [{
            'title': a.get_work_type_display(),
            'start': a.date.strftime('%Y-%m-%d'),
            'color': colors.get(a.work_type, '#6c757d')
        } for a in emp.attendance_set.all()]

        progress_percent = min(100, int((report['office_days'] / 8) * 100))

        data.append({
            'id': emp.id,
            'name': emp.full_name,
            'email': emp.email,
            'date_joined': emp.date_joined.strftime('%d/%m/%Y'),
            'office': report['office_days'],
            'total': report['total_days'],
            'is_ok': report['is_ok'],
            'debt': report['debt'],
            'today_status': today_status,
            'progress': progress_percent,
            'monthly_remaining': report['monthly_remaining'],
            'events_json': json.dumps(events_list)
        })

    current_month_atts = Attendance.objects.filter(
        date__year=today.year,
        date__month=today.month
    ).select_related('employee')

    today_atts = current_month_atts.filter(date=today)
    names_today_office = [a.employee.full_name for a in today_atts.filter(work_type='OFFICE')]
    names_today_remote = [a.employee.full_name for a in today_atts.filter(work_type='REMOTE')]
    names_today_leave = [a.employee.full_name for a in today_atts.filter(work_type__in=['LEAVE', 'SICK'])]

    names_month_office = sorted(list(set([a.employee.full_name for a in current_month_atts.filter(work_type='OFFICE')])))
    names_month_remote = sorted(list(set([a.employee.full_name for a in current_month_atts.filter(work_type='REMOTE')])))
    names_month_leave = sorted(list(set([a.employee.full_name for a in current_month_atts.filter(work_type__in=['LEAVE', 'SICK'])])))

    total_emp_count = len(data)
    presence_rate = (len(names_today_office) / total_emp_count * 100) if total_emp_count > 0 else 0

    stats_summary = {
        'today': {
            'office': len(names_today_office),
            'remote': len(names_today_remote),
            'leave': len(names_today_leave),
            'names_office': ", ".join(names_today_office) or "Κανένας",
            'names_remote': ", ".join(names_today_remote) or "Κανένας",
            'names_leave': ", ".join(names_today_leave) or "Κανένας",
        },
        'month': {
            'office': current_month_atts.filter(work_type='OFFICE').count(),
            'remote': current_month_atts.filter(work_type='REMOTE').count(),
            'leave': current_month_atts.filter(work_type__in=['LEAVE', 'SICK']).count(),
            'names_office': ", ".join(names_month_office) or "Κανένας",
            'names_remote': ", ".join(names_month_remote) or "Κανένας",
            'names_leave': ", ".join(names_month_leave) or "Κανένας",
        },
        'total_count': total_emp_count,
        'total_debt_sum': total_debt_accumulator,
        'presence_rate': round(presence_rate, 1)
    }

    return render(request, 'employees/manage.html', {
        'form': form,
        'employees': data,
        'stats': stats_summary,
        'holidays_js': json.dumps([d.strftime('%Y-%m-%d') for d in gr_holidays.keys()]),
        'holidays_events_json': json.dumps(holiday_events),
    })

@require_POST
def delete_employee(request, employee_id):
    emp = get_object_or_404(Employee, id=employee_id)
    name = emp.full_name
    emp.delete()
    messages.success(request, f"✅ Ο/Η {name} διαγράφηκε επιτυχώς.")
    return redirect('manage_employees')

@csrf_exempt
def update_attendance_ajax(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            emp_id = data.get('emp_id')
            new_type = data.get('work_type')
            date_str = data.get('date')

            selected_date = date.fromisoformat(date_str)

            if selected_date.weekday() in [5, 6]:
                return JsonResponse({'status': 'error', 'message': 'Δεν επιτρέπονται καταχωρήσεις Σαββατοκύριακα!'}, status=400)

            gr_holidays = holidays.Greece(years=selected_date.year)
            if selected_date in gr_holidays:
                return JsonResponse({'status': 'error', 'message': f'Η ημέρα είναι αργία: {gr_holidays.get(selected_date)}'}, status=400)

            if new_type == 'DELETE':
                Attendance.objects.filter(employee_id=emp_id, date=selected_date).delete()
                return JsonResponse({'status': 'success', 'action': 'deleted'})
            else:
                attendance, created = Attendance.objects.update_or_create(
                    employee_id=emp_id,
                    date=selected_date,
                    defaults={'work_type': new_type}
                )
                return JsonResponse({'status': 'success', 'action': 'created' if created else 'updated'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=400)

def export_attendance_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Μηνιαία Αναφορά Παρουσίας"

    headers = ['Ονοματεπώνυμο', 'Email', 'Γραφείο', 'Τηλεργασία', 'Άδειες/Ασθ.', 'Χρέος (Ημέρες)']
    ws.append(headers)

    header_fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    employees = Employee.objects.all()
    for emp in employees:
        report = emp.get_monthly_report()
        ws.append([
            emp.full_name,
            emp.email,
            report['office_days'],
            report['remote_days'],
            report['leave_days'],
            report['debt']
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="Attendance_Report.xlsx"'
    wb.save(response)
    return response

def export_attendance_pdf(request):
    try:
        pdfmetrics.registerFont(TTFont('GreekFont', 'DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('GreekFontBold', 'DejaVuSans-Bold.ttf'))
        font_main = 'GreekFont'
        font_bold = 'GreekFontBold'
    except:
        font_main = 'Helvetica'
        font_bold = 'Helvetica-Bold'

    today = date.today()
    month_name = today.strftime('%B %Y')

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.2 * cm, leftMargin=1.2 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        title=f"Αναφορά Παρουσιών {month_name}",
    )

    brand_red = colors.HexColor('#881337')
    light_red = colors.HexColor('#FFF1F2')
    red_text = colors.HexColor('#9F1239')
    green_bg = colors.HexColor('#F0FDF4')
    green_text = colors.HexColor('#166534')
    muted = colors.HexColor('#64748B')
    border_c = colors.HexColor('#E2E8F0')
    row_alt = colors.HexColor('#F8FAFC')

    title_style = ParagraphStyle('Title', fontName=font_bold, fontSize=22,
                                 textColor=brand_red, alignment=TA_LEFT, spaceAfter=10)

    sub_style = ParagraphStyle('Sub', fontName=font_main, fontSize=10,
                               textColor=muted, alignment=TA_LEFT, spaceAfter=2)

    name_style = ParagraphStyle('Name', fontName=font_main, fontSize=10,
                                textColor=colors.black, alignment=TA_LEFT, leftIndent=6)

    num_style = ParagraphStyle('Num', fontName=font_main, fontSize=10,
                               textColor=colors.black, alignment=TA_CENTER)

    ok_style = ParagraphStyle('OK', fontName=font_bold, fontSize=9,
                              textColor=green_text, alignment=TA_CENTER)
    debt_style = ParagraphStyle('Debt', fontName=font_bold, fontSize=9,
                                textColor=red_text, alignment=TA_CENTER)

    story = []

    story.append(Paragraph("PCS Attendance Management", title_style))
    story.append(Paragraph(f"Μηνιαία Αναφορά Παρουσιών — {month_name}", sub_style))
    story.append(Paragraph(f"Εκτυπώθηκε: {today.strftime('%d/%m/%Y')}", sub_style))
    story.append(Spacer(1, 0.6 * cm))

    employees = Employee.objects.all()
    total_debt = sum(emp.get_cumulative_debt() for emp in employees)
    emp_count = employees.count()
    ok_count = sum(1 for emp in employees if emp.get_cumulative_debt() == 0)

    summary_data = [
        ['Σύνολο Υπαλλήλων', 'Εντάξει', 'Με Χρέος', 'Συνολικό Χρέος'],
        [str(emp_count), str(ok_count), str(emp_count - ok_count), f"{total_debt} ημ."],
    ]
    summary_table = Table(summary_data, colWidths=[4.5 * cm] * 4)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), brand_red),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), font_bold),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.2, colors.white),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 1 * cm))

    col_widths = [6.5 * cm, 2.8 * cm, 2.8 * cm, 2.8 * cm, 3.5 * cm]
    headers = ['Ονοματεπώνυμο', 'Γραφείο', 'Τηλεργασία', 'Άδειες', 'Κατάσταση']
    table_data = [headers]

    for emp in employees:
        report = emp.get_monthly_report()
        display_name = emp.full_name.title()

        if report['is_ok']:
            status = Paragraph('<b>ΕΝΤΑΞΕΙ</b>', ok_style)
        else:
            status = Paragraph(f'<b>ΧΡΕΟΣ -{report["debt"]}</b>', debt_style)

        table_data.append([
            Paragraph(display_name, name_style),
            Paragraph(str(report['office_days']), num_style),
            Paragraph(str(report['remote_days']), num_style),
            Paragraph(str(report['leave_days']), num_style),
            status,
        ])

    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)

    row_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), brand_red),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), font_bold),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.2, border_c),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]

    for i in range(1, len(table_data)):
        bg = row_alt if i % 2 == 0 else colors.white
        row_styles.append(('BACKGROUND', (0, i), (-1, i), bg))

        emp_report = employees[i - 1].get_monthly_report()
        if emp_report['is_ok']:
            row_styles.append(('BACKGROUND', (4, i), (4, i), green_bg))
        else:
            row_styles.append(('BACKGROUND', (4, i), (4, i), light_red))

    main_table.setStyle(TableStyle(row_styles))
    story.append(main_table)

    story.append(Spacer(1, 0.8 * cm))
    story.append(Paragraph(
        f"* Η αναφορά αφορά τον μήνα {month_name}. Ελάχιστη απαίτηση: 8 ημέρες φυσικής παρουσίας στο γραφείο.",
        ParagraphStyle('Note', fontName=font_main, fontSize=8, textColor=muted, italic=True)
    ))

    doc.build(story)
    buffer.seek(0)

    filename = f"Attendance_Report_{today.strftime('%Y_%m')}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@csrf_exempt
def attendance_range_stats(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error'}, status=400)
    try:
        data       = json.loads(request.body)
        emp_id     = data.get('emp_id')
        date_from  = date.fromisoformat(data.get('date_from'))
        date_to    = date.fromisoformat(data.get('date_to'))

        if date_from > date_to:
            return JsonResponse({'status': 'error', 'message': 'Η αρχή πρέπει να είναι πριν το τέλος.'}, status=400)

        emp = get_object_or_404(Employee, id=emp_id)
        qs  = emp.attendance_set.filter(date__gte=date_from, date__lte=date_to)

        office = qs.filter(work_type='OFFICE').count()
        remote = qs.filter(work_type='REMOTE').count()
        leave  = qs.filter(work_type__in=['LEAVE', 'SICK']).count()
        total  = office + remote + leave

        return JsonResponse({
            'status': 'success',
            'office': office,
            'remote': remote,
            'leave':  leave,
            'total':  total,
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@csrf_exempt
def bulk_update_attendance(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            employee_ids = data.get('employee_ids', [])
            work_type = data.get('work_type')
            selected_date = date.today()

            if not employee_ids or not work_type:
                return JsonResponse({'status': 'error', 'message': 'Λείπουν δεδομένα'}, status=400)

            for emp_id in employee_ids:
                Attendance.objects.update_or_create(
                    employee_id=emp_id,
                    date=selected_date,
                    defaults={'work_type': work_type}
                )

            return JsonResponse({'status': 'success', 'count': len(employee_ids)})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error'}, status=400)